import base64
import glob
import os
import shutil

import fitz
from PIL import Image
from huaweicloudsdkcore.auth.credentials import BasicCredentials
from huaweicloudsdkcore.exceptions import exceptions
from huaweicloudsdkocr.v1 import *
from huaweicloudsdkocr.v1.region.ocr_region import OcrRegion
from openpyxl import Workbook, load_workbook


# def pdf2images(pdf_path, rotate, save_path):
#     if os.path.exists(save_path):
#         shutil.rmtree(save_path)
#     os.makedirs(save_path, exist_ok=True)
#     images = convert_from_path(pdf_path)
#     paths = []
#     for i, img in enumerate(images):
#         img = img.rotate(rotate, expand=True)
#         path = os.path.join(save_path, f'page-{i + 1:02}.jpg')
#         img.save(path, 'JPEG')
#         paths.append(path)
#     return paths


def pdf2images(pdf_path, rotate, save_path, start_idx, end_idx):
    if os.path.exists(save_path):
        shutil.rmtree(save_path)
    os.makedirs(save_path, exist_ok=True)

    zoom = 4  # to increase the resolution
    mat = fitz.Matrix(zoom, zoom)
    doc = fitz.open(pdf_path)
    start_idx -= 1
    if end_idx == 0:
        end_idx = len(doc) - 1
    else:
        end_idx -= 1
    paths = []
    for i, page in enumerate(doc):
        if start_idx <= i <= end_idx:
            pix = page.get_pixmap(matrix=mat)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            img = img.rotate(rotate, expand=True)
            path = os.path.join(save_path, f'page-{i + 1 :02}.jpg')
            img.save(path, 'JPEG')
            paths.append(path)
    return paths


def image2excel(image_path, excel_path, ak, sk, region):
    credentials = BasicCredentials(ak, sk)
    client = OcrClient.new_builder().with_credentials(credentials).with_region(OcrRegion.value_of(region)).build()
    with open(image_path, "rb") as f:
        image_base64 = base64.b64encode(f.read()).decode("utf-8")
    try:
        request = RecognizeGeneralTableRequest()
        request.body = GeneralTableRequestBody(return_excel=True, image=image_base64)
        response = client.recognize_general_table(request)
        excel_base64 = response.result.excel
        excel_data = base64.b64decode(excel_base64)
        with open(excel_path, "wb") as f:
            f.write(excel_data)
        return True, ""
    except exceptions.ClientRequestException as e:
        return False, e.error_msg


def combine_excels(excel_dir, save_path):
    file_list = glob.glob(excel_dir + "/*.xlsx")
    file_list.sort()
    dest_wb = Workbook()
    del dest_wb['Sheet']
    for i, file in enumerate(file_list):
        sheet_name = os.path.basename(file).split(".")[0]
        dest_wb.create_sheet(sheet_name)
        dest_ws = dest_wb[sheet_name]
        source_wb = load_workbook(file)
        source_sheet = source_wb.active
        for row in source_sheet.rows:
            for cell in row:
                dest_ws[cell.coordinate] = cell.value
    dest_wb.save(save_path)
